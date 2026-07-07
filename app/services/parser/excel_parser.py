from __future__ import annotations

import json
import re
import time
from collections import Counter, defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell, MergedCell
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from app.services.parser.img_parser import enrich_image_descriptions
from app.services.parser.paths import processing_subdir
from app.services.parser.word_parser import format_extracted_items


TABLE_JSON_MAX_CHARS = 800
NOTE_ROW_CELL_COUNT = 2
IMAGE_EXTENSIONS = {
    "png": ".png",
    "jpeg": ".jpg",
    "jpg": ".jpg",
    "gif": ".gif",
    "bmp": ".bmp",
    "tiff": ".tiff",
}


@dataclass
class ParseContext:
    source_path: Path
    image_dir: Path
    image_index: int = 0

    def next_image_path(self, image_format: str | None) -> Path:
        self.image_index += 1
        extension = IMAGE_EXTENSIONS.get(str(image_format or "").lower(), ".bin")
        return self.image_dir / f"image_{self.image_index:04d}{extension}"


def parse_excel_document(
    file_path: str | Path,
    *,
    image_analysis_workers: int = 3,
) -> list[dict[str, Any]]:
    """Extract xlsx sheets as sparse row JSON and enrich embedded images."""
    started_at = time.perf_counter()
    path = Path(file_path)
    if not path.exists():
        raise FileNotFoundError(f"Excel document not found: {path}")
    if path.suffix.lower() != ".xlsx":
        raise ValueError(f"Only .xlsx files are supported: {path}")
    if path.name.startswith("~$"):
        raise ValueError(f"Temporary Excel lock files are not supported: {path}")

    _log(f"start parsing: {path}")
    workbook = load_workbook(path, data_only=True)
    image_dir = processing_subdir(path, "img")
    image_dir.mkdir(parents=True, exist_ok=True)
    _remove_stale_bin_images(image_dir)

    stage_started_at = time.perf_counter()
    context = ParseContext(source_path=path, image_dir=image_dir)
    items = _extract_workbook_items(workbook.worksheets, context)
    _log_item_summary("extracted workbook items", items, stage_started_at)

    stage_started_at = time.perf_counter()
    _log("start image analysis")
    items = enrich_image_descriptions(items, document_title=path.stem, max_concurrency=image_analysis_workers)
    _log_item_summary("finished image analysis", items, stage_started_at)

    stage_started_at = time.perf_counter()
    output_path = write_items_to_txt(items, path)
    _log(f"wrote parsed txt: {output_path} ({time.perf_counter() - stage_started_at:.2f}s)")
    _log(f"finished parsing: {path.name} ({time.perf_counter() - started_at:.2f}s)")
    return items


def _extract_workbook_items(worksheets: list[Worksheet], context: ParseContext) -> list[dict[str, Any]]:
    items: list[dict[str, Any]] = [
        {
            "type": "paragraph",
            "style": "标题 1",
            "text": context.source_path.stem,
            "source": "workbook",
        }
    ]

    for worksheet in worksheets:
        _log(f"extracting sheet: {worksheet.title}")
        sheet_started_at = time.perf_counter()
        sheet_items = _extract_sheet_items(worksheet, context)
        items.extend(sheet_items)
        _log(
            f"finished sheet: {worksheet.title}, items={len(sheet_items)} "
            f"({time.perf_counter() - sheet_started_at:.2f}s)"
        )

    _log(f"scanned sheets={len(worksheets)}, images={context.image_index}")
    return items


def _extract_sheet_items(worksheet: Worksheet, context: ParseContext) -> list[dict[str, Any]]:
    items: list[dict[str, Any]] = [
        {
            "type": "paragraph",
            "style": "标题 2",
            "text": worksheet.title,
            "source": f"sheet:{worksheet.title}",
        }
    ]
    image_items_by_row = _extract_sheet_images(worksheet, context)

    first_data_row = _first_non_empty_row(worksheet)
    if first_data_row is None:
        return items

    header_row_index = first_data_row
    note_row = _special_note_row(worksheet, first_data_row)
    if note_row is not None:
        items.extend(
            _row_table_items(
                [note_row],
                source=f"sheet:{worksheet.title}:row:{first_data_row}",
                links=_row_links(worksheet, first_data_row),
            )
        )
        items.extend(image_items_by_row.pop(first_data_row, []))
        header_row_index = _next_non_empty_row(worksheet, first_data_row + 1) or first_data_row + 1

    headers = _unique_headers(_fill_empty_headers(_row_values(worksheet, header_row_index)))
    for image_item in image_items_by_row.pop(header_row_index, []):
        items.append(image_item)

    for row_index in range(header_row_index + 1, worksheet.max_row + 1):
        row = _row_object(worksheet, row_index, headers)
        if row:
            source = f"sheet:{worksheet.title}:row:{row_index}"
            items.extend(_row_table_items([row], source=source, links=_row_links(worksheet, row_index, headers)))
        items.extend(image_items_by_row.pop(row_index, []))

    for remaining_row in sorted(image_items_by_row):
        items.extend(image_items_by_row[remaining_row])

    return items


def _extract_sheet_images(worksheet: Worksheet, context: ParseContext) -> dict[int, list[dict[str, str]]]:
    images_by_row: dict[int, list[dict[str, str]]] = defaultdict(list)
    for image in getattr(worksheet, "_images", []):
        image_path = context.next_image_path(getattr(image, "format", None))
        image_path.write_bytes(image._data())
        row_index, column_index = _image_anchor_position(image)
        images_by_row[row_index].append(
            {
                "type": "image",
                "style": "图片",
                "source": f"sheet:{worksheet.title}:row:{row_index}:column:{column_index}",
                "text": str(image_path),
                "path": str(image_path),
            }
        )
    return dict(images_by_row)


def _image_anchor_position(image: Any) -> tuple[int, int]:
    anchor = getattr(image, "anchor", None)
    marker = getattr(anchor, "_from", None)
    if marker is None:
        return 1, 1
    return int(marker.row) + 1, int(marker.col) + 1


def _special_note_row(worksheet: Worksheet, row_index: int) -> dict[str, str] | None:
    values = _row_values(worksheet, row_index)
    meaningful_values = [value for value in values if _has_value(value)]
    if len(meaningful_values) != NOTE_ROW_CELL_COUNT:
        return None

    next_row_index = _next_non_empty_row(worksheet, row_index + 1)
    if next_row_index is None:
        return None
    if _non_empty_cell_count(_row_values(worksheet, next_row_index)) <= NOTE_ROW_CELL_COUNT:
        return None

    return {meaningful_values[0]: meaningful_values[1]}


def _row_object(worksheet: Worksheet, row_index: int, headers: list[str]) -> dict[str, str]:
    values = _row_values(worksheet, row_index)
    row: dict[str, str] = {}
    for index, header in enumerate(headers):
        if index >= len(values):
            break
        value = values[index]
        if _has_value(value):
            row[header] = value
    return row


def _row_values(worksheet: Worksheet, row_index: int) -> list[str]:
    return [_cell_value(worksheet.cell(row_index, column_index)) for column_index in range(1, worksheet.max_column + 1)]


def _row_links(worksheet: Worksheet, row_index: int, headers: list[str] | None = None) -> dict[str, str]:
    links: dict[str, str] = {}
    for column_index in range(1, worksheet.max_column + 1):
        cell = worksheet.cell(row_index, column_index)
        hyperlink = getattr(cell, "hyperlink", None)
        if hyperlink is None:
            continue

        url = str(getattr(hyperlink, "target", None) or getattr(hyperlink, "location", None) or "").strip()
        if not url:
            continue

        header = headers[column_index - 1] if headers and column_index <= len(headers) else ""
        description = (
            _normalize_cell_value(getattr(cell, "value", None))
            or str(getattr(hyperlink, "display", "") or "").strip()
            or header
            or url
        )
        links[_unique_link_description(description, links)] = url
    return links


def _unique_link_description(description: str, links: dict[str, str]) -> str:
    if description not in links:
        return description

    index = 2
    while f"{description} {index}" in links:
        index += 1
    return f"{description} {index}"


def _cell_value(cell: Cell | MergedCell) -> str:
    value = cell.value
    if value is None and isinstance(cell, MergedCell):
        value = _merged_cell_parent_value(cell)
    return _normalize_cell_value(value)


def _merged_cell_parent_value(cell: MergedCell) -> Any:
    worksheet = cell.parent
    coordinate = cell.coordinate
    for merged_range in worksheet.merged_cells.ranges:
        if coordinate in merged_range:
            if merged_range.max_row == merged_range.min_row:
                return None
            return worksheet.cell(merged_range.min_row, merged_range.min_col).value
    return None


def _normalize_cell_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        text = str(int(value))
    else:
        text = str(value)
    lines = [re.sub(r"[ \t\u3000]+", " ", line).strip() for line in text.splitlines()]
    return "\n".join(line for line in lines if line).strip()


def _first_non_empty_row(worksheet: Worksheet) -> int | None:
    return _next_non_empty_row(worksheet, 1)


def _next_non_empty_row(worksheet: Worksheet, start_row: int) -> int | None:
    for row_index in range(start_row, worksheet.max_row + 1):
        if _non_empty_cell_count(_row_values(worksheet, row_index)) > 0:
            return row_index
    return None


def _non_empty_cell_count(values: list[str]) -> int:
    return sum(1 for value in values if _has_value(value))


def _has_value(value: str) -> bool:
    return bool(str(value).strip())


def _fill_empty_headers(headers: list[str]) -> list[str]:
    filled_headers: list[str] = []
    last_header = ""
    for index, header in enumerate(headers, start=1):
        normalized = header.strip()
        if normalized:
            last_header = normalized
            filled_headers.append(normalized)
        elif last_header:
            filled_headers.append(last_header)
        else:
            filled_headers.append(f"column_{get_column_letter(index)}")
    return filled_headers


def _unique_headers(headers: list[str]) -> list[str]:
    counts: dict[str, int] = {}
    unique_headers: list[str] = []
    for header in headers:
        counts[header] = counts.get(header, 0) + 1
        unique_headers.append(header if counts[header] == 1 else f"{header}_{counts[header]}")
    return unique_headers


def _row_table_items(
    rows: list[dict[str, str]],
    *,
    source: str,
    links: dict[str, str] | None = None,
) -> list[dict[str, Any]]:
    if not rows:
        return []
    item: dict[str, Any] = {
        "type": "table",
        "style": "表格",
        "source": source,
        "text": json.dumps(rows, ensure_ascii=False, separators=(",", ":")),
    }
    if links:
        item["links"] = dict(links)
    return [item]


def write_items_to_txt(items: list[dict[str, Any]], source_file: str | Path) -> Path:
    source_path = Path(source_file)
    txt_dir = processing_subdir(source_path, "txt")
    txt_dir.mkdir(parents=True, exist_ok=True)

    output_path = txt_dir / f"{source_path.stem}.txt"
    output_path.write_text(format_extracted_items(items), encoding="utf-8")
    return output_path


def _remove_stale_bin_images(image_dir: Path) -> None:
    for path in image_dir.glob("*.bin"):
        path.unlink(missing_ok=True)


def _log(message: str) -> None:
    print(f"[excel_parser] {message}", flush=True)


def _log_item_summary(message: str, items: list[dict[str, Any]], started_at: float) -> None:
    counts = Counter(item.get("type", "unknown") for item in items)
    type_summary = ", ".join(f"{item_type}={count}" for item_type, count in sorted(counts.items()))
    _log(f"{message}: total={len(items)} ({type_summary or 'no items'}) ({time.perf_counter() - started_at:.2f}s)")


if __name__ == "__main__":
    target_file = Path("data") / "raw" / "普通表格" / "迈拓供应商信息全览表.xlsx"
    extracted_items = parse_excel_document(target_file)
    txt_path = Path("data") / "processing" / target_file.stem / "txt" / f"{target_file.stem}.txt"

    print(f"File: {target_file}")
    print(f"Extracted text blocks: {len(extracted_items)}")
    print(f"Text output: {txt_path}")
